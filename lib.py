import re
import os

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)


FOLDER = "cases"
OUTPUT = "result"
NUM_FORMAT = "#,##0"
RUB_FORMAT = '#,##0.00 ₽'

import json
import pandas as pd
from dataclasses import dataclass
import threading

def subtract_df(df1, df2):
    COUNT_COL = 12

    if df1.shape[1] < COUNT_COL or df2.shape[1] < COUNT_COL:
        raise ValueError("DataFrames should have at least COUNT_COL columns")

    df1['temp_key'] = df1.iloc[:, :COUNT_COL].astype(str).apply(lambda x: '_'.join(x), axis=1)
    df2['temp_key'] = df2.iloc[:, :COUNT_COL].astype(str).apply(lambda x: '_'.join(x), axis=1)

    df1 = df1.drop_duplicates(subset='temp_key')
    df2 = df2.drop_duplicates(subset='temp_key')

    merged_df = pd.merge(df1, df2[['temp_key']], on='temp_key', how='left', indicator=True)

    filtered_df = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge', 'temp_key'])

    return filtered_df



def parse_conditions_to_code(conditions):
    python_code = conditions.replace("AND", "&").replace("OR", "|").replace("NOT", "~")

    python_code = python_code.replace('"Площадка"', 'df_orig["Площадка"]')
    python_code = python_code.replace('"Тип используемых прав" ', 'df_orig["Тип используемых прав"]')
    python_code = python_code.replace('"Территория"', 'df_orig["Территория"]')
    python_code = python_code.replace('"Тип контента"', 'df_orig["Тип контента"]')
    python_code = python_code.replace('"Вид использования  контента"', 'df_orig["Вид использования  контента"]')
    python_code = python_code.replace('"Исполнитель"', 'df_orig["Исполнитель"]')
    python_code = python_code.replace('"Название трека"', 'df_orig["Название трека"]')
    python_code = python_code.replace('"Название альбома"', 'df_orig["Название альбома"]')
    python_code = python_code.replace('"Автор слов"', 'df_orig["Автор слов"]')
    python_code = python_code.replace('"Автор музыки"', 'df_orig["Автор музыки"]')
    python_code = python_code.replace('"ISRC контента"', 'df_orig["ISRC контента"]')
    python_code = python_code.replace('"UPC альбома"', 'df_orig["UPC альбома"]')
    python_code = python_code.replace('"Копирайт"', 'df_orig["Копирайт"]')

    python_code = re.sub(r'\s+', ' ', python_code).strip()

    python_code = f'df_orig[{python_code}]'

    return python_code


def calc_sum(df):
    res = 0
    for index, row in df.iterrows():
        res += (row.iloc[c["Вознаграждение_1"]] + row.iloc[c["Вознаграждение_2"]])
    return round(res, 2)


def calc_sum_before(df):
    res = 0
    for index, row in df.iterrows():
        res += (row.iloc[c["Сумма_1"]] + row.iloc[c["Сумма_2"]])
    return round(res, 2)


@dataclass
class Data:
    license: int
    track: int
    name: str

def parse_filename(name: str) -> Data:
    name = name.replace(".txt", "")
    d = name.split(",")

    name = d[0].replace("name=", "")
    license = int(d[1].replace("license=", ""))
    track = int(d[2].replace("track=", ""))

    return Data(license=license, track=track, name=name)

with open("config.json", "r", encoding="utf-8") as f:
    c = json.loads(f.read())


report_name = c["filename"].split('.')[0]

import numpy as np 
def read_df(c, result):
    df_orig = pd.read_excel(c["filename"], header=c["header_index_from_zero"])

    col_count = len(df_orig.columns)
    if  col_count > c["col_count"]:
        df_orig = df_orig.iloc[:, :c["col_count"]-col_count]

    df_orig["UPC альбома"] = df_orig["UPC альбома"].replace(np.nan, 0).astype('longlong').astype(str)
    result.append(df_orig)

    print(df_orig["UPC альбома"])

result = []
threading.Thread(target=read_df, args=(c,result), daemon=True).start()

all_case_dfs = list()


def save_rest_df():
    global all_case_dfs, result
    if not result:
        return 

    subtract_df(result[0], pd.concat(all_case_dfs)).to_excel("невошедшее.xlsx", index=False)



def get_percent(data: Data):
    return ((data.license*data.track)/10000)


def change_data_in_columns(df, a, b):
    сумма_средств_1 = 18
    сумма_средств_2 = 19
    вознаграждение_1 = 21
    вознаграждение_2 = 22
    доля_лицензиара = 20
    итого_вознаграждения = 23

    df.iloc[:, сумма_средств_1] = df.iloc[:, вознаграждение_1]
    df.iloc[:, сумма_средств_2] = df.iloc[:, вознаграждение_2]
    df.iloc[:, итого_вознаграждения] = (df.iloc[:, вознаграждение_1] + df.iloc[:, вознаграждение_2]) * (a * b) / 10000
    df.iloc[:, вознаграждение_1] = df.iloc[:, вознаграждение_1] * (a * b) / 10000
    df.iloc[:, вознаграждение_2] = df.iloc[:, вознаграждение_2] * (a * b) / 10000
    df.iloc[:, доля_лицензиара] = f"{a}%"


    return df

def run(case):
    global result

    if not result:
        return None, None, None

    if case == "":
        return 0,0,0

    df_orig = result[0]
    case_dfs = []

    case = case.replace(f"{FOLDER}{os.path.sep}", "")

    case_path = os.path.join(FOLDER, case)
    sum_all = 0
    sum_after = 0

    if os.path.isfile(case_path):
        with open(case_path, "r", encoding="utf-8") as f:
            conditions = f.read()
        code_output = parse_conditions_to_code(conditions)
        data = parse_filename(case)
        df = change_data_in_columns(eval(code_output), data.license, data.track)

        case_dfs.append(df)
        all_case_dfs.append(df)

        sum_all += calc_sum(df)
        sum_after += calc_sum_before(df)
    else:
        for subcase in os.listdir(case_path):
            r = 0
            with open(os.path.join(case_path, subcase), "r", encoding="utf-8") as f:
                conditions = f.read()
            code_output = parse_conditions_to_code(conditions)
            data = parse_filename(subcase)

            df = change_data_in_columns(eval(code_output), data.license, data.track)

            case_dfs.append(df)
            all_case_dfs.append(df)

            sum_all += calc_sum(df)

            sum_after += calc_sum_before(df)
        data.name = case

    res_df = pd.concat(case_dfs).drop_duplicates()

    column_index = 23
    total_sum = res_df.iloc[:,  column_index].sum()


    last_row_data = {
        16: "Итого: ",
        17: res_df.iloc[:,  17].sum(),
        18: res_df.iloc[:,  18].sum(),
        19: res_df.iloc[:,  19].sum(),
        21: res_df.iloc[:,  21].sum(),
        22: res_df.iloc[:,  22].sum(),
        23: res_df.iloc[:,  23].sum(),
    }

    new_row = pd.DataFrame([last_row_data[i] if i in last_row_data else None for i in range(len(res_df.columns))]).T
    new_row.columns = res_df.columns

    res_df = pd.concat([res_df, new_row], ignore_index=True)

    if not os.path.exists(OUTPUT):
        os.mkdir(OUTPUT)

    if "" in case:
        case_name = case.split(',')[0].replace("name=","").replace("cases\\", "")
        excel_filepath = os.path.join(OUTPUT, f"{case_name} {report_name}.xlsx", )
    else:
        excel_filepath = os.path.join(OUTPUT, f"{case} {report_name}.xlsx", )


    res_df = res_df.drop(res_df.columns[-1], axis=1)
    res_df = res_df.drop(res_df.columns[0], axis=1)
    res_df = res_df.reset_index(drop=True)
    res_df.to_excel(excel_filepath, index=False, sheet_name="Детализированный отчёт")

    grouped_df = res_df.groupby('Площадка').agg({
        'Количество загрузок/прослушиваний': 'sum',
        'Итого вознаграждение ЛИЦЕНЗИАРА в руб., без НДС': 'sum'
    })

    grouped_df.rename(columns={
        "Количество загрузок/прослушиваний": "Количество загрузок / прослушиваний",
        "Итого вознаграждение ЛИЦЕНЗИАРА в руб., без НДС": "Итого вознаграждение лицензиара, руб.",
    }, inplace=True)

    last_row_data_gr = {
        0: grouped_df.iloc[:,  0].sum(),
        1: grouped_df.iloc[:,  1].sum(),
    }

    new_row_gr = pd.DataFrame([last_row_data_gr[i] if i in last_row_data_gr else None for i in range(len(grouped_df.columns))]).T
    new_row_gr.columns = grouped_df.columns
    grouped_df = pd.concat([grouped_df, new_row_gr], ignore_index=False)


    with pd.ExcelWriter(excel_filepath, engine='openpyxl', mode='a') as writer:
        grouped_df.to_excel(writer, sheet_name='Сводный отчёт')

    import openpyxl
    from openpyxl.styles import Border, Side, Alignment, Font, numbers

    workbook = openpyxl.load_workbook(excel_filepath)

    bold_font = Font(bold=True)

    for sheet_name in workbook.sheetnames:
        workbook[sheet_name].sheet_view.zoomScale = 80
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        if sheet_name == "Сводный отчёт":
            for row in workbook[sheet_name].iter_rows(
                min_row=1, 
                max_row=workbook[sheet_name].max_row, 
                min_col=1, 
                max_col=workbook[sheet_name].max_column
            ):
                for cell in row:
                    if row.index == workbook[sheet_name].max_row - 1:
                        cell.font = bold_font
                    cell.border = thin_border

        cols_to_format = ['L', 'M']

        for col in cols_to_format:
            for cell in workbook[sheet_name][col]:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00

        cols_to_format_rub = ['R', 'S', 'U', 'V', 'W']

        for col in cols_to_format_rub:
            for cell in workbook[sheet_name][col]:
                cell.number_format = RUB_FORMAT

        for cell in workbook[sheet_name]['Q']:
            cell.number_format = NUM_FORMAT

        if sheet_name == "Детализированный отчёт":
            last_row = workbook[sheet_name].max_row

            for cell in workbook[sheet_name][last_row]:
                if cell.column_letter != 'Q':
                    cell.number_format = RUB_FORMAT
        else:
            last_row = workbook[sheet_name].max_row

            workbook[sheet_name][last_row][0].value = "Итого: "
            workbook[sheet_name][last_row][1].number_format = NUM_FORMAT
            workbook[sheet_name][last_row][2].number_format = RUB_FORMAT

            for i in range(1, workbook[sheet_name].max_row):
                workbook[sheet_name][i][1].number_format = NUM_FORMAT
                workbook[sheet_name][i][2].number_format = RUB_FORMAT

    workbook.save(excel_filepath)

    return len(res_df), round(sum_after, 2), round(sum_all, 2)
